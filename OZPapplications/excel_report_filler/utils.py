import re
from fuzzywuzzy import fuzz
from openpyxl.utils import get_column_letter

class Utils:
    """
    Содержит вспомогательные функции для различных операций в приложении.
    """
    def __init__(self, config_manager=None):
        self.config_manager = config_manager

    def extract_address_from_filename(self, filename):
        """
        Извлекает адрес из имени файла с помощью регулярного выражения из конфига.
        """
        if not self.config_manager:
            # Fallback для тестирования без ConfigManager
            pattern = r'\(([^)]+)\)'
            # print("Warning: ConfigManager not provided to Utils. Using default regex.")
        else:
            pattern = self.config_manager.get('Regex', 'address_extraction_pattern')

        match = re.search(pattern, filename)
        if match:
            return match.group(1).strip()
        return None

    def fuzzy_match(self, search_text, candidates_list, threshold=80):
        """
        Находит наилучшее нечёткое совпадение `search_text` в `candidates_list`.
        Возвращает (совпадение, score), если score выше порога, иначе (None, 0).
        """
        if not candidates_list:
            return None, 0

        best_match = None
        highest_score = 0

        for candidate in candidates_list:
            score = fuzz.ratio(search_text.lower(), candidate.lower())
            if score > highest_score:
                highest_score = score
                best_match = candidate

        if highest_score >= threshold:
            return best_match, highest_score
        return None, 0

    def find_cell_by_keywords(self, worksheet, keywords, search_range=None):
        """
        Находит ячейку, содержащую одно из ключевых слов (регистронезависимо).
        `keywords` - список строк.
        `search_range` - кортеж (min_row, min_col, max_row, max_col) для ограничения поиска.
        Возвращает (row, col) найденной ячейки или None.
        """
        # Преобразуем ключевые слова в нижний регистр для регистронезависимого поиска
        lower_keywords = [k.lower() for k in keywords]

        if search_range:
            min_row, min_col, max_row, max_col = search_range
        else:
            min_row, min_col, max_row, max_col = 1, 1, worksheet.max_row, worksheet.max_column

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None and isinstance(cell_value, str):
                    for keyword in lower_keywords:
                        if keyword in cell_value.lower():
                            return row, col # Возвращаем 1-based индексы
        return None

    def find_value_cell(self, worksheet, start_row, start_col, max_search_distance=5):
        """
        Пытается найти подходящую ячейку для записи значения,
        начиная от start_cell и двигаясь вправо, затем вниз.
        Учитывает объединенные ячейки.
        Возвращает (row, col) ячейки для записи или None.
        """
        # Проверяем ячейки справа
        for col_offset in range(1, max_search_distance + 1):
            target_col = start_col + col_offset
            if target_col > worksheet.max_column:
                break
            cell = worksheet.cell(row=start_row, column=target_col)
            if self._is_suitable_for_value(cell, worksheet):
                return cell.row, cell.column

        # Если справа не нашли, проверяем ячейки снизу
        for row_offset in range(1, max_search_distance + 1):
            target_row = start_row + row_offset
            if target_row > worksheet.max_row:
                break
            cell = worksheet.cell(row=target_row, column=start_col)
            if self._is_suitable_for_value(cell, worksheet):
                return cell.row, cell.column

        return None

    def _is_suitable_for_value(self, cell, worksheet):
        """
        Проверяет, подходит ли ячейка для записи значения.
        Ячейка подходит, если она пуста или содержит только пробелы,
        или если она является частью объединенной ячейки,
        и эта объединенная ячейка пуста или содержит только пробелы.
        """
        # Проверяем, является ли ячейка частью объединенной области
        for merged_cell_range in worksheet.merged_cells.ranges:
            if cell.coordinate in merged_cell_range:
                # Если это объединенная ячейка, проверяем ее верхнюю левую ячейку
                # т.к. только она хранит значение для объединенной области
                min_col_letter = get_column_letter(merged_cell_range.min_col)
                top_left_cell_coord = f"{min_col_letter}{merged_cell_range.min_row}"
                top_left_cell = worksheet[top_left_cell_coord]
                return top_left_cell.value is None or str(top_left_cell.value).strip() == ''

        # Если ячейка не объединена, проверяем ее значение
        return cell.value is None or str(cell.value).strip() == ''

    def get_cell_value(self, worksheet, row, col):
        """Безопасно получает значение ячейки, учитывая объединенные ячейки."""
        cell = worksheet.cell(row=row, column=col)
        for merged_range in worksheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Если ячейка находится в объединенном диапазоне, возвращаем значение из верхней левой ячейки
                return worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        return cell.value

    def get_boolean_from_text(self, text):
        """Преобразует текстовое значение в булево (для наличия газа)."""
        if text is None:
            return False
        text_lower = str(text).strip().lower()
        if text_lower in ('да', 'true', 'есть', 'yes', '+'):
            return True
        return False