import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
import re
import datetime


class ReportProcessor:
    """
    Обрабатывает Excel-отчёты: сканирует, извлекает данные, заполняет поля,
    вставляет строки для членов комиссии (если требуется) и генерирует отчёты о выполнении.
    """
    def __init__(self, config_manager, commission_manager, utils, log_callback=None):
        self.config_manager = config_manager
        self.commission_manager = commission_manager
        self.utils = utils
        self.log_message = log_callback if log_callback else print
        self.report_files = [] # Список найденных файлов отчетов

        # Настраиваемые параметры для поиска газа в отчете
        self.gas_detection_keywords = [
            k.strip() for k in self.config_manager.get('Regex', 'gas_detection_keywords').split(',')
        ]
        self.gas_cell_offset_x = int(self.config_manager.get('Regex', 'gas_detection_cell_offset_x'))
        self.gas_cell_offset_y = int(self.config_manager.get('Regex', 'gas_detection_cell_offset_y'))

        # Для хранения ручных сопоставлений полей (пока не реализовано в UI)
        try:
            self.manual_field_mappings = self.config_manager.get('FieldMapping', 'manual_mappings')
        except Exception:
            self.manual_field_mappings = {}
    def _detect_gas_in_report(self, report_path):
        """
        Определяет, есть ли газ в отчёте.
        Заглушка: всегда возвращает False.
        Реализуйте свою логику, если нужно.
        """
        # TODO: Реализовать определение наличия газа по содержимому отчёта
        return False

    def scan_reports(self, reports_folder):
        """Сканирует указанную папку на наличие файлов отчётов Excel."""
        self.report_files = []
        if not os.path.isdir(reports_folder):
            self.log_message(f"Папка с отчётами не найдена: {reports_folder}", level="error")
            return []

        for root, _, files in os.walk(reports_folder):
            for file in files:
                if file.lower().endswith(('.xlsx', '.xls')):
                    full_path = os.path.join(root, file)
                    self.report_files.append(full_path)
        self.log_message(f"Найдено {len(self.report_files)} файлов отчётов в {reports_folder}", level="info")
        return self.report_files

    def process_all_reports(self, reports_folder, data_folder, output_folder, update_progress_callback=None):
        """Обрабатывает все найденные отчёты."""
        if not self.report_files:
            self.log_message("Нет отчётов для обработки. Сначала просканируйте папку.", level="warning")
            return

        results = []
        total_reports = len(self.report_files)
        for i, report_path in enumerate(self.report_files):
            file_name = os.path.basename(report_path)
            self.log_message(f"Обработка отчёта {i+1}/{total_reports}: {file_name}", level="info")
            if update_progress_callback:
                update_progress_callback(i + 1, total_reports, file_name)

            report_result = self.process_single_report(report_path, data_folder, output_folder)
            results.append(report_result)

        self._generate_processing_report(output_folder, results)
        self.log_message("Обработка всех отчётов завершена.", level="success")
        return results

    def process_single_report(self, report_path, data_folder, output_folder):
        """
        Обрабатывает один файл отчёта: извлекает данные, заполняет поля,
        вставляет строки и сохраняет.
        Возвращает словарь с результатом обработки.
        """
        file_name = os.path.basename(report_path)
        address = self.utils.extract_address_from_filename(file_name)
        if not address:
            self.log_message(f"Не удалось извлечь адрес из имени файла '{file_name}'. Пропускаю.", level="error")
            return {
                "file": file_name,
                "address": "Неизвестен",
                "status": "Ошибка",
                "message": "Не удалось извлечь адрес из имени файла",
                "filled_fields": 0,
                "missing_data_fields": []
            }

        data_file_path = self._find_data_file_for_address(data_folder, address)
        if not data_file_path:
            self.log_message(f"Не найден файл данных для адреса '{address}'. Пропускаю '{file_name}'.", level="warning")
            return {
                "file": file_name,
                "address": address,
                "status": "Ошибка",
                "message": "Не найден файл данных для адреса",
                "filled_fields": 0,
                "missing_data_fields": []
            }

        try:
            data_from_file = self._read_data_file(data_file_path)
        except Exception as e:
            self.log_message(f"Ошибка чтения файла данных {data_file_path}: {e}. Пропускаю '{file_name}'.", level="error")
            return {
                "file": file_name,
                "address": address,
                "status": "Ошибка",
                "message": f"Ошибка чтения файла данных: {e}",
                "filled_fields": 0,
                "missing_data_fields": []
            }

        # Теперь объединяем данные из файла и данные комиссии
        full_data_for_report = data_from_file.copy()

        try:
            workbook = load_workbook(report_path)
            sheet = workbook.active # Или выбрать конкретный лист, если нужно
            self.log_message(f"Открыт отчёт: {file_name}", level="info")

            # 1. Определение наличия газа в отчёте
            has_gas_in_report = self._detect_gas_in_report(sheet)
            self.log_message(f"Для '{address}' обнаружено газоснабжение: {'Да' if has_gas_in_report else 'Нет'}", level="info")

            # 2. Получение состава комиссии
            commission_composition = self.commission_manager.get_commission_composition(address, has_gas_in_report)
            if commission_composition:
                self.log_message(f"Состав комиссии для '{address}' ({'Газ' if has_gas_in_report else 'Без газа'}) найден.", level="info")
                full_data_for_report.update(commission_composition) # Добавляем данные комиссии к общим данным
            else:
                self.log_message(f"Не удалось найти состав комиссии для '{address}' ({'Газ' if has_gas_in_report else 'Без газа'}).", level="warning")


            filled_count = 0
            missing_fields = []
            matched_cells = {} # Для избежания повторного заполнения одной ячейки

            # Если есть "Ресурсник" и газа не было, возможно нужно добавить строки
            # Эта логика должна быть более сложной и зависеть от шаблона отчета
            # Для примера, предположим, что ресурсник всегда добавляется последним
            # А также где в отчете искать место для вставки.
            # Пока оставим простой поиск и запись
            if has_gas_in_report and 'Ресурсник' in full_data_for_report:
                # Находим строку, после которой нужно вставить нового члена комиссии
                # Например, ищем "Член комиссии", и после него вставляем "Ресурсника"
                # Это очень сильно зависит от шаблона отчета!
                # Для примера, ищем "Член комиссии"
                member_row_col = self.utils.find_cell_by_keywords(sheet, ["Член комиссии", "Член"])
                if member_row_col:
                    insert_row = member_row_col[0] + 1 # Вставляем после члена комиссии
                    self._insert_rows(sheet, insert_row, 1) # Вставляем одну строку
                    self.log_message(f"Вставлена строка для ресурсника после строки {insert_row-1}.", level="info")
                else:
                    self.log_message("Не удалось найти 'Член комиссии' для вставки строки ресурсника. Заполнение будет произведено в существующие поля.", level="warning")


            # Ищем и заполняем поля в отчете
            for data_field, data_value in full_data_for_report.items():
                found_cell_coords = None
                # Сначала пытаемся найти по точному совпадению или ручному маппингу
                if data_field in self.manual_field_mappings:
                    target_coord = self.manual_field_mappings[data_field]
                    if re.match(r"^[A-Z]+\d+$", target_coord): # Проверка формата A1
                        found_cell_coords = (sheet[target_coord].row, sheet[target_coord].column)
                else:
                    # Ищем поле нечётко
                    # Предполагаем, что ключевые слова для поиска полей могут быть в заголовках столбцов или рядом
                    # Проходим по всем ячейкам листа в поисках совпадения
                    for row_idx in range(1, sheet.max_row + 1):
                        for col_idx in range(1, sheet.max_column + 1):
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            cell_value = str(cell.value).strip() if cell.value is not None else ""

                            matched_keyword, score = self.utils.fuzzy_match(data_field, [cell_value])
                            if score >= 85: # Высокий порог для прямых совпадений
                                # Нашли потенциальное поле, теперь ищем ячейку для значения
                                value_row, value_col = self.utils.find_value_cell(sheet, cell.row, cell.column)
                                if value_row and value_col:
                                    # Проверим, что мы не заполняем ту же ячейку снова, если поле уже было найдено
                                    if (value_row, value_col) not in matched_cells.values():
                                        found_cell_coords = (value_row, value_col)
                                        break
                        if found_cell_coords:
                            break

                if found_cell_coords:
                    row, col = found_cell_coords
                    current_cell_value = self.utils.get_cell_value(sheet, row, col)
                    if current_cell_value is None or str(current_cell_value).strip() == '':
                        # Записываем значение
                        sheet.cell(row=row, column=col, value=data_value)
                        matched_cells[data_field] = (row, col)
                        filled_count += 1
                        # self.log_message(f"  Заполнено поле '{data_field}' в ячейке {get_column_letter(col)}{row} значением '{data_value}'", level="debug")
                    else:
                        self.log_message(f"  Ячейка {get_column_letter(col)}{row} для поля '{data_field}' уже содержит значение: '{current_cell_value}'. Пропускаю.", level="warning")
                else:
                    missing_fields.append(data_field)
                    self.log_message(f"  Не найдено подходящее место для заполнения поля '{data_field}'", level="warning")


            # Сохранение заполненного отчёта
            output_file_name = f"{os.path.splitext(file_name)[0]}_FILLED.xlsx"
            output_path = os.path.join(output_folder, output_file_name)
            workbook.save(output_path)
            self.log_message(f"Отчёт '{file_name}' успешно заполнен и сохранён как '{output_file_name}'", level="success")

            return {
                "file": file_name,
                "address": address,
                "status": "Успешно",
                "message": "Отчёт успешно заполнен",
                "filled_fields": filled_count,
                "missing_data_fields": missing_fields,
                "has_gas": has_gas_in_report
            }

        except Exception as e:
            self.log_message(f"Критическая ошибка при обработке отчёта '{file_name}': {e}", level="error")
            return {
                "file": file_name,
                "address": address,
                "status": "Ошибка",
                "message": f"Критическая ошибка: {e}",
                "filled_fields": 0,
                "missing_data_fields": list(full_data_for_report.keys()) # Все поля могли быть незаполнены
            }

    def _find_data_file_for_address(self, base_data_folder, address):
        """
        Ищет файл данных, содержащий в названии нужную подстроку, в base_data_folder и подпапках.
        """
        target_substring = "Объемы выполненных работ по подготовке объекта к эксплуатации"
        for root, dirs, files in os.walk(base_data_folder):
            for f in files:
                if (
                    f.lower().endswith(('.xlsx', '.xls', '.csv'))
                    and target_substring.lower() in f.lower()
                ):
                    return os.path.join(root, f)
        return None

    def _read_data_file(self, file_path):
        """
        Читает данные из Excel или CSV файла и возвращает их в виде словаря.
        Предполагается, что данные находятся в первом листе, и формат:
        первый столбец - название поля, второй столбец - значение.
        """
    
        if file_path.lower().endswith('.csv'):
            df = pd.read_csv(file_path, header=None, encoding='utf-8')
        else:
            df = pd.read_excel(file_path, engine="openpyxl")
    
        data_dict = {}
        for index, row in df.iterrows():
            if len(row) >= 2 and pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
                key = str(row.iloc[0]).strip()
                value = str(row.iloc[1]).strip()
                data_dict[key] = value
        return data_dict
        """
        Пытается определить наличие газоснабжения в отчете по ключевым словам.
        Возвращает True, если найдено, False иначе.
        """
        for keyword in self.gas_detection_keywords:
            found_pos = self.utils.find_cell_by_keywords(worksheet, [keyword])
            if found_pos:
                # Если ключевое слово найдено, проверяем смещенную ячейку
                target_row = found_pos[0] + self.gas_cell_offset_y
                target_col = found_pos[1] + self.gas_cell_offset_x
                
                if target_row > 0 and target_col > 0: # Убедимся, что индексы валидны
                    cell_value = self.utils.get_cell_value(worksheet, target_row, target_col)
                    if self.utils.get_boolean_from_text(cell_value):
                        return True
        return False

    def _insert_rows(self, worksheet, start_row, num_rows):
        """
        Вставляет `num_rows` пустых строк, начиная со `start_row`.
        Корректирует объединенные ячейки, которые пересекаются с точкой вставки.
        """
        # Сдвигаем все строки ниже start_row
        worksheet.insert_rows(start_row, num_rows)

        # Корректируем объединенные ячейки
        # Важно: openpyxl.insert_rows обычно сам корректирует объединенные ячейки,
        # но сложные случаи могут потребовать ручной обработки.
        # Для базового случая insert_rows должен сработать.
        # Дополнительная проверка на всякий случай
        for merged_range_str in list(worksheet.merged_cells.ranges): # Создаем копию списка, т.к. будем изменять
            merged_range = CellRange(merged_range_str)
            if merged_range.min_row >= start_row:
                # Если объединенная ячейка находится полностью ниже точки вставки, просто сдвигаем
                new_min_row = merged_range.min_row + num_rows
                new_max_row = merged_range.max_row + num_rows
                new_range = CellRange(min_col=merged_range.min_col, max_col=merged_range.max_col,
                                      min_row=new_min_row, max_row=new_max_row)
                worksheet.merged_cells.remove(merged_range_str)
                worksheet.merged_cells.add(str(new_range))
            elif merged_range.min_row < start_row <= merged_range.max_row:
                # Если точка вставки внутри объединенной ячейки, расширяем ее
                new_max_row = merged_range.max_row + num_rows
                new_range = CellRange(min_col=merged_range.min_col, max_col=merged_range.max_col,
                                      min_row=merged_range.min_row, max_row=new_max_row)
                worksheet.merged_cells.remove(merged_range_str)
                worksheet.merged_cells.add(str(new_range))


    def _generate_processing_report(self, output_folder, results):
        """Генерирует Excel-отчёт о результатах обработки."""
        report_data = []
        for res in results:
            report_data.append({
                "Файл отчёта": res["file"],
                "Адрес": res["address"],
                "Статус": res["status"],
                "Сообщение": res["message"],
                "Заполнено полей": res["filled_fields"],
                "Незаполненные поля (из данных)": ", ".join(res["missing_data_fields"]),
                "Наличие газа в отчёте": "Да" if res.get("has_gas") else "Нет"
            })

        df = pd.DataFrame(report_data)
        report_file_name = f"Отчёт_об_обработке_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = os.path.join(output_folder, report_file_name)

        try:
            df.to_excel(report_path, index=False)
            self.log_message(f"Отчёт об обработке успешно создан: {report_path}", level="success")
        except Exception as e:
            self.log_message(f"Ошибка при создании отчёта об обработке: {e}", level="error")